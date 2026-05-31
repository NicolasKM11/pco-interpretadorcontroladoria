import re
from io import BytesIO

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from PIL import Image

try:
    import pytesseract
except Exception:
    pytesseract = None

st.set_page_config(page_title="Interpretador PCO", page_icon="📊", layout="wide")

CONTAS_PADRAO = [
    "Empréstimos", "Desp. Financeiras", "Faturamento", "CPV", "Despesas Operacionais",
    "Lucro Operacional", "Lucro Líquido", "CDG", "NCG", "Tesouraria", "Contas a Receber",
    "Estoque MP", "Estoque PA", "Ativo Operacional", "Fornecedores", "Tributos", "Outros", "Passivo Operacional"
]

SINONIMOS = {
    "Empréstimos": ["emprést", "emprest", "emp.", "emprestimos"],
    "Desp. Financeiras": ["desp. fin", "desp fin", "despesas financeiras", "d.fin", "desp financeiras"],
    "Faturamento": ["faturamento", "fat"],
    "CPV": ["c p v", "cpv", "c.p.v"],
    "Despesas Operacionais": ["despesas operacionais", "desp oper", "despesas", "desp. oper"],
    "Lucro Operacional": ["lucro op", "lucro operacional", "lo"],
    "Lucro Líquido": ["lucro liq", "lucro líq", "lucro liquido", "ll"],
    "CDG": ["cdg"],
    "NCG": ["ncg"],
    "Tesouraria": ["tesouraria", "tes"],
    "Contas a Receber": ["contas a receber", "cr"],
    "Estoque MP": ["estoque mp", "est. mp", "est mp"],
    "Estoque PA": ["estoque pa", "est. pa", "est pa", "produto acabado"],
    "Ativo Operacional": ["ativo op", "ativo operacional", "ao"],
    "Fornecedores": ["fornecedor", "forn", "forncedores"],
    "Tributos": ["tribut"],
    "Outros": ["outros"],
    "Passivo Operacional": ["passivo op", "passivo operacional", "po"],
}

PERGUNTAS_PADRAO = {
    "Quantidade vendida": [
        ("Resultado Contábil — por que o Δ Faturamento é superior ao Δ CPV?",
         "O faturamento aumentou diretamente pelo crescimento da quantidade vendida. O CPV também aumenta, mas tende a crescer em proporção menor porque parte dos custos é fixa e passa a ser diluída em um volume maior de produção. Por isso, o aumento percentual do faturamento fica superior ao aumento percentual do CPV, melhorando a margem e o resultado operacional."),
        ("Resultado Patrimonial [NCG] — o que justifica o aumento em Fornecedores [S103]?",
         "O aumento em fornecedores ocorre porque, para sustentar o maior volume vendido, a empresa precisa produzir mais e comprar mais matéria-prima. Como parte dessas compras é realizada a prazo, o saldo de fornecedores no passivo operacional aumenta."),
        ("Resultado Financeiro — que variável contribuiu positivamente para reduzir empréstimos?",
         "A variável positiva é o aumento do faturamento e dos recebimentos. Como parte das vendas entra no caixa, a empresa gera mais recursos próprios e passa a depender menos de empréstimos."),
        ("Resultado Financeiro — que variável contribuiu negativamente para reduzir empréstimos?",
         "O fator negativo é que vender mais exige mais produção, mais compras, mais estoques e mais contas a receber. Esse crescimento da operação aumenta a necessidade de capital de giro e consome parte da melhora financeira."),
        ("Resultado Contábil — por que o Δ CPV é inferior ao Δ faturamento?",
         "O CPV cresce menos que o faturamento porque a empresa passa a produzir e vender mais unidades usando a mesma estrutura fixa. Assim, o custo fixo é diluído em mais peças, reduzindo o custo unitário e fazendo o CPV crescer proporcionalmente menos."),
        ("Resultado Patrimonial [NCG] — por que aumenta Contas a Receber?",
         "O contas a receber aumenta porque o faturamento cresceu e parte das vendas é feita a prazo. Portanto, uma parcela maior das vendas fica registrada para recebimento futuro."),
    ],
    "Preço de venda": [
        ("Resultado Patrimonial [NCG] — o que justifica o aumento no CR/Contas a Receber [S98]?",
         "O contas a receber aumenta porque o preço de venda maior eleva o faturamento. Como parte das vendas ocorre a prazo, o valor financeiro ainda não recebido pela empresa também aumenta."),
        ("Resultado Contábil — explique a variação nas Despesas Operacionais [O87]",
         "As despesas operacionais variam porque algumas contas, como comissões, marketing/publicidade e outras despesas variáveis, acompanham o faturamento. Quando o preço aumenta e a receita cresce, essas despesas também podem crescer, mesmo sem alteração relevante na quantidade produzida."),
        ("Resultado Contábil — por que o CPV fica igual ou quase igual?",
         "Quando a premissa altera apenas o preço de venda, a quantidade produzida e o consumo de matéria-prima permanecem praticamente iguais. Por isso, o CPV tende a ficar igual ou variar pouco."),
        ("Resultado Contábil — por que o Δ LL pode ser muito maior que no aumento da quantidade?",
         "O aumento de preço melhora diretamente a margem, porque eleva o faturamento sem exigir aumento proporcional de produção, matéria-prima e estrutura. Por isso, o ganho no lucro líquido pode ser mais expressivo do que no cenário de aumento da quantidade vendida."),
        ("Resultado Contábil — por que LO é igual ao LL?",
         "O lucro operacional pode ser igual ao lucro líquido quando a empresa ainda permanece em prejuízo. Nesse caso, não há incidência de imposto sobre lucro, então o resultado operacional e o resultado líquido ficam iguais."),
        ("Resultado Financeiro — por que o impacto financeiro é mais expressivo que no aumento de quantidade?",
         "O reajuste de preço aumenta o faturamento sem exigir o mesmo crescimento de produção, compras e desembolsos. Assim, há maior entrada de recursos com menor pressão operacional, melhorando mais os empréstimos e as despesas financeiras."),
    ],
    "Produção / Estoque PA": [
        ("Resultado Contábil — por que o impacto no resultado contábil pode não ser significativo [O89]?",
         "O impacto pode não ser significativo porque a premissa altera principalmente o nível de estoque e a forma de produção, sem mexer diretamente no faturamento. Assim, o efeito aparece mais no custo unitário e no capital de giro do que na receita."),
        ("Resultado Contábil — por que a readequação do estoque PA pode aumentar o CPV?",
         "A readequação do estoque de produtos acabados pode reduzir a produção em alguns períodos. Com menor produção, os custos fixos são distribuídos em menos unidades, elevando o custo unitário e, consequentemente, o CPV."),
        ("Resultado Patrimonial [NCG] — por que houve redução na NCG?",
         "A NCG pode reduzir porque a diminuição do estoque de produto acabado reduz o ativo operacional. Quando o ativo operacional cai mais do que o passivo operacional, a necessidade de capital de giro diminui."),
        ("Produção pontual — por que o custo unitário varia muito?",
         "Na produção pontual, a empresa não produz a mesma quantidade todos os meses. Nos meses de menor produção, os custos fixos são absorvidos por menos unidades, elevando o custo unitário. Nos meses de maior produção, ocorre o efeito contrário."),
        ("Produção constante — por que gera mais previsibilidade?",
         "A produção constante distribui a produção ao longo dos meses de forma mais linear. Isso facilita planejamento de compras, pessoal e capacidade produtiva, mas pode aumentar estoques quando a venda não acompanha o mesmo ritmo."),
    ],
    "Compra / Estoque MP": [
        ("Resultado Contábil — por que a readequação do estoque MP pode ter impacto positivo?",
         "A readequação do estoque de matéria-prima pode reduzir compras, armazenagem e necessidade de financiamento. Isso melhora o resultado principalmente quando diminui desembolsos, empréstimos ou despesas financeiras."),
        ("Resultado Contábil — se a empresa não estivesse endividada, qual seria o impacto?",
         "Se a empresa não estivesse endividada, a redução de compras teria menor impacto no resultado contábil, porque não haveria uma economia relevante de despesas financeiras. O benefício ficaria mais concentrado no caixa e na necessidade de capital de giro."),
        ("Resultado Patrimonial [NCG] — por que houve redução em Fornecedores [S103]?",
         "A conta fornecedores reduz quando a empresa passa a comprar menos matéria-prima ou reduz o nível de estoque de MP. Comprando menos a prazo, o saldo de obrigações com fornecedores também diminui."),
        ("Compra pontual — por que pode piorar a NCG?",
         "A compra pontual concentra compras em determinados meses e deixa a operação mais exposta a variações de preço, prazo e necessidade de caixa. Isso pode elevar fornecedores, estoques ou desembolsos em períodos específicos, piorando a NCG."),
        ("Compra constante — por que reduz risco operacional?",
         "A compra constante deixa as aquisições mais previsíveis e reduz a exposição a oscilações pontuais de preço e fornecimento. Por outro lado, pode manter estoques maiores e exigir mais capital parado em matéria-prima."),
    ],
    "Preço dos insumos": [
        ("Resultado Financeiro — que variável contribuiu positivamente para a redução dos empréstimos?",
         "A redução no preço dos insumos diminui os desembolsos com compras e melhora a geração de caixa. Com menor necessidade de recursos para financiar a operação, a empresa depende menos de empréstimos."),
        ("Resultado Contábil — por que houve redução no CPV?",
         "O CPV reduz porque o preço de aquisição dos insumos caiu. Como matéria-prima é componente relevante do custo de produção, a redução do custo dos insumos diminui o custo total dos produtos vendidos."),
        ("Resultado Patrimonial [NCG] — por que reduziu Estoque de PA [S100]?",
         "O estoque de produtos acabados pode reduzir em valor porque os produtos passam a carregar um custo unitário menor. Mesmo com quantidade parecida, o valor contábil do estoque diminui quando o custo de produção cai."),
    ],
    "PMC / Prazo de recebimento": [
        ("Resultado Patrimonial [NCG] — por que o CR aumenta quando o prazo aumenta?",
         "Quando a empresa concede mais prazo ao cliente, uma parcela maior das vendas fica para recebimento futuro. Assim, mesmo com faturamento parecido, o contas a receber aumenta e pressiona o ativo operacional e a NCG."),
        ("Resultado Financeiro — por que prazo maior piora empréstimos?",
         "Com prazo maior de recebimento, o dinheiro demora mais para entrar no caixa. A empresa precisa financiar a operação por mais tempo, aumentando a dependência de empréstimos e as despesas financeiras."),
        ("Resultado Patrimonial [NCG] — por que prazo menor melhora a NCG?",
         "Prazo menor acelera os recebimentos e reduz contas a receber. Como o ativo operacional diminui, a necessidade de capital de giro tende a melhorar."),
    ],
    "PMO / Prazo de pagamento": [
        ("Resultado Patrimonial [NCG] — por que aumentar PMO melhora a NCG?",
         "Ao aumentar o prazo médio obtido com fornecedores, a empresa demora mais para pagar suas compras. Isso aumenta o passivo operacional e reduz a necessidade de capital de giro, pois parte da operação passa a ser financiada pelos fornecedores."),
        ("Resultado Financeiro — por que pagar fornecedores mais tarde reduz empréstimos?",
         "Pagando fornecedores em prazo maior, a empresa preserva caixa por mais tempo. Com isso, precisa captar menos empréstimos para financiar a operação."),
        ("Resultado Patrimonial [NCG] — por que reduzir PMO piora a NCG?",
         "Reduzir o prazo de pagamento antecipa desembolsos. O passivo operacional diminui e a empresa precisa usar mais capital próprio ou empréstimos para sustentar a operação."),
    ],
    "Despesas / MKT / Comissão / Transporte": [
        ("Resultado Contábil — por que as despesas operacionais aumentaram?",
         "As despesas operacionais aumentam quando há crescimento de vendas, faturamento ou parâmetros como marketing, comissão e transporte. Comissão e publicidade costumam acompanhar o faturamento, enquanto transporte acompanha o volume entregue."),
        ("Resultado Contábil — por que transporte aumenta?",
         "O transporte aumenta quando há mais unidades vendidas/entregues ou quando o custo unitário de transporte é reajustado. Mesmo não sendo sempre percentual do faturamento, ele acompanha o volume físico movimentado."),
        ("Resultado Contábil — por que comissão aumenta?",
         "A comissão aumenta porque normalmente é calculada como percentual do faturamento ou das vendas. Se a empresa fatura mais ou aumenta a taxa de comissão, a despesa com comissão cresce."),
        ("Resultado Contábil — por que MKT/Publicidade aumenta?",
         "Marketing e publicidade aumentam quando são definidos como percentual do faturamento ou quando há reforço de investimento comercial. Com faturamento maior, o valor absoluto da despesa também aumenta."),
    ],
    "Pessoal / RH": [
        ("Resultado Contábil — por que reajuste salarial piora o resultado?",
         "O reajuste salarial aumenta os custos de pessoal e encargos. Como essa despesa não gera aumento automático de faturamento, ela tende a elevar custos/despesas e reduzir o lucro operacional."),
        ("Resultado Contábil — por que admissões podem aumentar CPV/despesas?",
         "Admissões aumentam salários, encargos e custos ligados à produção ou administração. Se não houver aumento proporcional de produtividade ou vendas, o resultado contábil piora."),
        ("Resultado Financeiro — por que pessoal impacta caixa?",
         "Gastos com pessoal geram desembolsos recorrentes. Quando salários e encargos aumentam, a empresa precisa de mais caixa para cumprir essas obrigações, podendo elevar a pressão sobre empréstimos."),
    ],
    "Outra": [
        ("Interpretação geral",
         "A análise deve observar o efeito dominó da premissa sobre RF, RC e RP/NCG. Primeiro avalie faturamento, CPV e despesas; depois empréstimos e despesas financeiras; por fim, contas a receber, estoques, fornecedores, CDG, NCG e tesouraria."),
    ]
}



# Base construída com TODOS os materiais enviados na conversa:
# - PDFs de elaboração de premissas;
# - Cenário 1;
# - Cenário 2;
# - prints de questões/provas;
# - anotações de aula;
# - planilha P2 ATT.xlsm.
# A ideia é não decorar uma planilha específica, e sim responder pela lógica: premissa -> RF -> RC -> RP/NCG.
MATERIAL_BASE = {
    "Elaboração de premissas": [
        "Vendas envolve quantidades, preços de venda e política comercial/PMC.",
        "Produção envolve quantidade produzida, estoque e metodologia constante ou pontual.",
        "Compras envolve quantidade, preço de aquisição, estoque de matéria-prima, metodologia de compra e PMO.",
        "Pessoal envolve admissões, férias, dispensas, remunerações e reajustes.",
        "Custos e despesas envolvem MKT como percentual do faturamento, transporte e despesas variáveis.",
        "O processo de controle parte dos resultados críticos, analisa causas, define alternativas, redige premissas, testa e mede a contribuição de cada premissa."
    ],
    "Roteiro oficial de análise": [
        "As premissas devem ser testadas individualmente.",
        "Antes de nova premissa, retornar aos valores iniciais/originais quando o exercício pedir.",
        "A análise deve quantificar consequências no resultado financeiro, contábil e patrimonial/NCG.",
        "Além de quantificar, questionar se a premissa é factível.",
        "O objetivo é entender o efeito dominó entre operações, caixa, lucro e capital de giro."
    ],
    "Resultado Financeiro RF": [
        "Verificar se a empresa possui passivo financeiro.",
        "Analisar impacto da premissa nos empréstimos projetados.",
        "Analisar o impacto dos empréstimos nas despesas financeiras.",
        "Empréstimos e despesas financeiras normalmente andam juntos: menos empréstimos reduzem juros; mais empréstimos aumentam juros."
    ],
    "Resultado Contábil RC": [
        "Analisar faturamento, CPV, despesas totais, lucro operacional e lucro líquido.",
        "Se aumenta quantidade vendida, o faturamento cresce e o CPV costuma crescer menos proporcionalmente pela diluição de custos fixos.",
        "Se aumenta preço sem aumentar quantidade, o faturamento cresce sem impacto direto relevante na produção/CPV.",
        "Despesas operacionais podem aumentar por transporte, comissões, publicidade/MKT e despesas variáveis ligadas ao faturamento ou volume."
    ],
    "Resultado Patrimonial RP/NCG": [
        "O impacto no CDG decorre principalmente do lucro/prejuízo acumulado.",
        "Para explicar NCG, avaliar Ativo Operacional e Passivo Operacional.",
        "Ativo Operacional inclui contas a receber, estoque de matéria-prima e estoque de produto acabado.",
        "Passivo Operacional inclui fornecedores, tributos, salários/encargos e outros passivos operacionais.",
        "NCG = AO - PO. Se AO cresce mais que PO, a NCG piora/aumenta. Se AO cai ou PO aumenta, a NCG melhora/diminui.",
        "Tesouraria é afetada pela relação entre CDG e NCG; quando o CDG melhora mais que a NCG, a tesouraria melhora."
    ],
    "Questões recorrentes": [
        "Por que faturamento cresce mais que CPV? Diluição de custos fixos.",
        "Por que fornecedores aumentam? Mais compras a prazo para sustentar produção/vendas.",
        "Por que contas a receber aumenta? Maior faturamento e/ou maior prazo concedido.",
        "Por que despesas operacionais mudam? MKT, comissões, transporte e despesas variáveis acompanham faturamento/volume ou premissa definida.",
        "Por que LO = LL? Quando ainda há prejuízo, não há imposto sobre lucro.",
        "Por que reduzir estoque melhora NCG? Menor ativo operacional imobilizado.",
        "Por que reduzir produção pode piorar CPV? Menor diluição de custo fixo aumenta custo unitário.",
        "Por que aumentar PMO melhora NCG? Fornecedores financiam a operação por mais tempo.",
        "Por que aumentar PMC piora NCG? O caixa demora mais para entrar e o CR aumenta."
    ]
}

PERGUNTAS_EXTRAS_DOCUMENTOS = {
    "Geral / Roteiro oficial": [
        ("Como responder qualquer pergunta da prova?",
         "Use o roteiro oficial: primeiro identifique a premissa alterada; depois explique o efeito no Resultado Financeiro (empréstimos e despesas financeiras), no Resultado Contábil (faturamento, CPV, despesas, lucro operacional e lucro líquido) e no Resultado Patrimonial/NCG (AO, PO, NCG, CDG e tesouraria). A resposta deve mostrar causa e consequência, não apenas repetir o número."),
        ("Por que as premissas devem ser analisadas individualmente?",
         "Porque o objetivo é isolar o efeito de cada decisão. Se várias premissas forem alteradas ao mesmo tempo sem controle, não dá para saber qual variável causou a melhora ou piora em RF, RC ou RP/NCG."),
        ("O que significa efeito dominó na simulação?",
         "É a sequência de impactos gerada por uma premissa. Por exemplo: vender mais aumenta faturamento, exige mais produção e compras, altera CPV e despesas, muda contas a receber/fornecedores, afeta NCG, empréstimos, despesa financeira e tesouraria."),
        ("Como avaliar se a premissa é factível?",
         "Além de observar a melhora nos números, é necessário avaliar se a empresa realmente consegue implementar a decisão: vender mais, reajustar preço sem perder mercado, reduzir estoque sem faltar produto, negociar prazos com fornecedores ou cortar custos sem prejudicar a operação.")
    ],
    "Quantidade vendida": [
        ("Por que vender mais pode melhorar empréstimos e ao mesmo tempo aumentar NCG?",
         "Vender mais melhora o caixa e o resultado, porque aumenta o faturamento. Porém, também aumenta a operação: mais produção, mais compras, mais contas a receber e mais estoques. Por isso, os empréstimos podem cair pela maior geração de caixa, enquanto a NCG aumenta porque a empresa precisa financiar uma operação maior."),
        ("Por que o aumento de vendas gera aumento de tributos?",
         "Porque tributos sobre vendas acompanham o faturamento. Se a empresa vende mais ou vende por preço maior, a base de cálculo dos tributos cresce e a conta de tributos a recolher também aumenta."),
        ("Por que o estoque de PA pode mudar em valor mesmo sem grande mudança em quantidade?",
         "Porque o estoque é registrado pelo valor de custo. Se o custo unitário cai pela diluição dos custos fixos, o valor contábil do estoque pode diminuir ou crescer menos, mesmo que a quantidade física não tenha mudado muito.")
    ],
    "Preço de venda": [
        ("Por que aumento de preço pode ser melhor que aumento de quantidade no resultado contábil?",
         "Porque o aumento de preço eleva o faturamento sem exigir, necessariamente, mais produção, mais matéria-prima e mais estrutura. Assim, uma parcela maior da receita adicional vira melhoria de margem e lucro."),
        ("Por que aumentar preço pode piorar NCG, mesmo sendo bom para lucro?",
         "Porque o faturamento maior aumenta o valor das vendas a prazo. Isso eleva contas a receber e o ativo operacional. Portanto, a NCG pode aumentar, mesmo que o cenário seja positivo para lucro e caixa."),
        ("Qual cuidado com aumento de preço?",
         "É preciso avaliar se o mercado aceita o novo preço. A premissa pode melhorar os números, mas não ser factível se reduzir demanda, gerar perda de clientes ou deixar o produto menos competitivo.")
    ],
    "Produção / Estoque PA": [
        ("Por que produção pontual reduz estoque mas pode aumentar custo unitário?",
         "A produção pontual acompanha a demanda e pode reduzir estoques. Porém, nos meses de baixa produção, os custos fixos são divididos por menos unidades, elevando o custo unitário."),
        ("Por que reduzir estoque PA melhora o financeiro mas pode piorar o contábil?",
         "Melhora o financeiro porque reduz capital parado em estoque e diminui necessidade de compras/financiamento. Pode piorar o contábil porque produzir menos em alguns meses aumenta o custo unitário e pode elevar CPV."),
        ("Quando reduzir estoque é ruim?",
         "Quando a redução força queda grande de produção, aumenta demais o custo unitário ou gera risco de falta de produto para vender. Nesse caso, a melhora de caixa pode não compensar a perda operacional.")
    ],
    "Compra / Estoque MP": [
        ("Por que reduzir estoque de MP melhora NCG?",
         "Porque estoque de matéria-prima faz parte do ativo operacional. Ao reduzir esse estoque, a empresa diminui recursos parados na operação, reduzindo a necessidade de capital de giro."),
        ("Por que comprar menos reduz fornecedores?",
         "Porque fornecedores representam compras realizadas a prazo. Se a empresa compra menos matéria-prima, o saldo a pagar aos fornecedores tende a diminuir."),
        ("Qual risco da compra pontual?",
         "A compra pontual pode deixar a empresa exposta a variações de preço, disponibilidade de matéria-prima e necessidade de caixa em meses específicos. Ela pode melhorar alguns indicadores, mas aumenta risco operacional.")
    ],
    "PMC / PMO": [
        ("Por que reduzir PMC melhora a tesouraria?",
         "Porque a empresa recebe mais rápido dos clientes. Isso reduz contas a receber, melhora o giro de caixa, diminui NCG e pode reduzir empréstimos."),
        ("Por que aumentar PMO pode ser positivo?",
         "Porque a empresa ganha mais prazo para pagar fornecedores. Isso aumenta o passivo operacional saudável e financia parte da operação sem recorrer tanto a empréstimos bancários."),
        ("Qual cuidado ao mexer em PMC e PMO?",
         "PMC menor pode dificultar vendas se o cliente não aceitar prazo curto. PMO maior depende de negociação com fornecedores e pode gerar perda de desconto ou piora na relação comercial.")
    ]
}

# Incorpora as perguntas extras ao banco usado pelo app.
for _premissa_extra, _itens_extra in PERGUNTAS_EXTRAS_DOCUMENTOS.items():
    PERGUNTAS_PADRAO.setdefault(_premissa_extra, [])
    PERGUNTAS_PADRAO[_premissa_extra].extend(_itens_extra)


TODAS_PERGUNTAS = []
for _prem, _lista in PERGUNTAS_PADRAO.items():
    for _titulo, _texto in _lista:
        TODAS_PERGUNTAS.append({"premissa": _prem, "titulo": _titulo, "resposta": _texto})


def normaliza(x):
    return re.sub(r"\s+", " ", str(x).strip().lower())


def to_float(x):
    if x is None or x == "":
        return None
    if isinstance(x, (int, float)):
        return float(x)
    s = str(x).replace("R$", "").replace("%", "").strip()
    s = s.replace("−", "-").replace("–", "-")
    # pt-BR: 1.234.567,89
    s = s.replace(".", "").replace(",", ".")
    try:
        return float(s)
    except Exception:
        return None


def detectar_conta(texto):
    t = normaliza(texto)
    if not t:
        return None
    for conta, termos in SINONIMOS.items():
        if any(term in t for term in termos):
            return conta
    return None


def montar_df(registros):
    if not registros:
        return pd.DataFrame(columns=["Conta", "Original", "Novo Valor", "Variação", "%"])
    df = pd.DataFrame(registros)
    df = df.dropna(subset=["Conta"])
    df = df.drop_duplicates(subset=["Conta"], keep="last")
    ordem = {c: i for i, c in enumerate(CONTAS_PADRAO)}
    df["ordem"] = df["Conta"].map(ordem).fillna(999)
    df = df.sort_values("ordem").drop(columns="ordem")
    for col in ["Original", "Novo Valor"]:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0.0)
    df["Variação"] = df["Novo Valor"] - df["Original"]
    df["%"] = df.apply(lambda r: ((r["Variação"] / abs(r["Original"])) * 100) if r["Original"] else 0, axis=1)
    return df


def ler_planilha(uploaded):
    data = uploaded.read()
    wb = load_workbook(BytesIO(data), data_only=True, read_only=True)
    candidatos = []
    # procura linhas com nome da conta e dois números próximos à direita
    for ws in wb.worksheets:
        for linha_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            vals = list(row)
            for i, v in enumerate(vals):
                conta = detectar_conta(v)
                if conta:
                    nums = []
                    for j in range(i + 1, min(i + 14, len(vals))):
                        n = to_float(vals[j])
                        if n is not None:
                            nums.append(n)
                    if len(nums) >= 2:
                        candidatos.append({
                            "Conta": conta,
                            "Original": nums[0],
                            "Novo Valor": nums[1],
                            "Aba": ws.title,
                            "Linha": linha_idx,
                        })
    return montar_df(candidatos)


def ocr_imagem(uploaded):
    if pytesseract is None:
        return ""
    img = Image.open(uploaded).convert("RGB")
    # aumenta a imagem para melhorar OCR em prints de planilha
    w, h = img.size
    if w < 1800:
        img = img.resize((w * 2, h * 2))
    try:
        texto = pytesseract.image_to_string(img, lang="por")
    except Exception:
        texto = pytesseract.image_to_string(img)
    return texto


def extrair_df_texto(texto):
    registros = []
    linhas = texto.splitlines()
    num_re = re.compile(r"-?\d{1,3}(?:\.\d{3})*,\d{2}|-?\d+[,.]\d+|-?\d+")
    for linha in linhas:
        conta = detectar_conta(linha)
        if not conta:
            continue
        numeros = [to_float(x.group()) for x in num_re.finditer(linha)]
        numeros = [n for n in numeros if n is not None]
        if len(numeros) >= 2:
            registros.append({"Conta": conta, "Original": numeros[0], "Novo Valor": numeros[1]})
    return montar_df(registros)


def sinal(v):
    if pd.isna(v) or abs(v) < 0.01:
        return "não teve variação relevante"
    return "aumentou" if v > 0 else "diminuiu"


def val(df, conta, campo="Variação"):
    if df is None or df.empty or campo not in df.columns:
        return 0.0
    linha = df[df["Conta"] == conta]
    if linha.empty:
        return 0.0
    try:
        return float(linha.iloc[0][campo])
    except Exception:
        return 0.0


def pct(df, conta):
    return val(df, conta, "%")


def frase_moeda(v):
    return f"R$ {v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")



def resposta_por_pergunta(pergunta, premissa, df):
    """Escolhe uma resposta-base pelo texto da pergunta, mesmo se a pergunta vier em imagem/OCR."""
    q = normaliza(pergunta)
    if not q:
        return []

    regras = [
        (["faturamento", "superior", "cpv"], "Resultado Contábil — por que o Δ Faturamento é superior ao Δ CPV?"),
        (["cpv", "inferior", "faturamento"], "Resultado Contábil — por que o Δ CPV é inferior ao Δ faturamento?"),
        (["fornecedor", "aumento"], "Resultado Patrimonial [NCG] — o que justifica o aumento em Fornecedores [S103]?"),
        (["fornecedor", "redu"], "Resultado Patrimonial [NCG] — por que houve redução em Fornecedores [S103]?"),
        (["contas a receber"], "Resultado Patrimonial [NCG] — o que justifica o aumento no CR/Contas a Receber [S98]?"),
        (["cr"], "Resultado Patrimonial [NCG] — o que justifica o aumento no CR/Contas a Receber [S98]?"),
        (["despesas operacionais"], "Resultado Contábil — explique a variação nas Despesas Operacionais [O87]"),
        (["desp", "oper"], "Resultado Contábil — explique a variação nas Despesas Operacionais [O87]"),
        (["impacto", "resultado contábil", "não", "significativo"], "Resultado Contábil — por que o impacto no resultado contábil pode não ser significativo [O89]?"),
        (["estoque", "pa", "cpv"], "Resultado Contábil — por que a readequação do estoque PA pode aumentar o CPV?"),
        (["lucro operacional", "lucro líquido"], "Resultado Contábil — por que LO é igual ao LL?"),
        (["lo", "ll"], "Resultado Contábil — por que LO é igual ao LL?"),
        (["emprést", "positivamente"], "Resultado Financeiro — que variável contribuiu positivamente para reduzir empréstimos?"),
        (["empr", "positivamente"], "Resultado Financeiro — que variável contribuiu positivamente para reduzir empréstimos?"),
        (["emprést", "negativamente"], "Resultado Financeiro — que variável contribuiu negativamente para reduzir empréstimos?"),
        (["empr", "negativamente"], "Resultado Financeiro — que variável contribuiu negativamente para reduzir empréstimos?"),
        (["preço", "insumo", "cpv"], "Resultado Contábil — por que houve redução no CPV?"),
        (["preços", "aquisição", "estoque"], "Resultado Patrimonial [NCG] — por que reduziu Estoque de PA [S100]?"),
        (["pmo"], "Resultado Patrimonial [NCG] — por que aumentar PMO melhora a NCG?"),
        (["prazo", "pagamento"], "Resultado Patrimonial [NCG] — por que aumentar PMO melhora a NCG?"),
        (["pmc"], "Resultado Patrimonial [NCG] — por que o CR aumenta quando o prazo aumenta?"),
        (["prazo", "recebimento"], "Resultado Patrimonial [NCG] — por que o CR aumenta quando o prazo aumenta?"),
        (["transporte"], "Resultado Contábil — por que transporte aumenta?"),
        (["comiss"], "Resultado Contábil — por que comissão aumenta?"),
        (["mkt"], "Resultado Contábil — por que MKT/Publicidade aumenta?"),
        (["publicidade"], "Resultado Contábil — por que MKT/Publicidade aumenta?"),
        (["sal"], "Resultado Contábil — por que reajuste salarial piora o resultado?"),
        (["rh"], "Resultado Contábil — por que reajuste salarial piora o resultado?"),
    ]

    # procura primeiro em todas as perguntas pelo título exato mapeado
    escolhidas = []
    for termos, titulo_alvo in regras:
        if all(t in q for t in termos):
            for item in TODAS_PERGUNTAS:
                if item["titulo"] == titulo_alvo:
                    escolhidas.append((item["titulo"], adaptar_resposta_valores(item["resposta"], df)))
                    break
    if escolhidas:
        return escolhidas

    # fallback: pontuação simples por palavras comuns
    melhor = []
    for item in TODAS_PERGUNTAS:
        titulo_norm = normaliza(item["titulo"])
        score = sum(1 for token in q.split() if len(token) > 3 and token in titulo_norm)
        if item["premissa"] == premissa:
            score += 1
        if score > 0:
            melhor.append((score, item))
    melhor.sort(key=lambda x: x[0], reverse=True)
    if melhor:
        item = melhor[0][1]
        return [(item["titulo"], adaptar_resposta_valores(item["resposta"], df))]

    return [("Resposta geral para a pergunta", adaptar_resposta_valores(PERGUNTAS_PADRAO["Outra"][0][1], df))]


def adaptar_resposta_valores(texto, df):
    """Adiciona uma frase curta com os valores encontrados para deixar a resposta específica do cenário."""
    if df is None or df.empty:
        return texto
    fat, cpv, desp = val(df, "Faturamento"), val(df, "CPV"), val(df, "Despesas Operacionais")
    emp, dfin = val(df, "Empréstimos"), val(df, "Desp. Financeiras")
    ncg, cdg, tes = val(df, "NCG"), val(df, "CDG"), val(df, "Tesouraria")
    detalhes = []
    for nome, variacao in [("faturamento", fat), ("CPV", cpv), ("despesas operacionais", desp), ("empréstimos", emp), ("despesas financeiras", dfin), ("NCG", ncg), ("CDG", cdg), ("tesouraria", tes)]:
        if abs(variacao) > 0.01:
            detalhes.append(f"{nome} {sinal(variacao)} ({frase_moeda(variacao)})")
    if detalhes:
        return texto + "\n\nPelos valores lidos na planilha/imagem: " + "; ".join(detalhes[:6]) + "."
    return texto

def gerar_respostas(df, premissa):
    fat, cpv, desp = val(df, "Faturamento"), val(df, "CPV"), val(df, "Despesas Operacionais")
    emp, dfin = val(df, "Empréstimos"), val(df, "Desp. Financeiras")
    cr, forn = val(df, "Contas a Receber"), val(df, "Fornecedores")
    ao, po = val(df, "Ativo Operacional"), val(df, "Passivo Operacional")
    ncg, cdg, tes = val(df, "NCG"), val(df, "CDG"), val(df, "Tesouraria")
    pfat, pcpv = pct(df, "Faturamento"), pct(df, "CPV")

    respostas = list(PERGUNTAS_PADRAO.get(premissa, PERGUNTAS_PADRAO["Outra"]))

    conclusao = []
    if fat > 0:
        conclusao.append("o faturamento aumentou, indicando maior geração de receita")
    elif fat < 0:
        conclusao.append("o faturamento diminuiu, reduzindo a geração de receita")
    if fat > 0 and cpv > 0 and abs(pfat) > abs(pcpv):
        conclusao.append("o CPV cresceu proporcionalmente menos que o faturamento, sinalizando ganho de diluição/margem")
    if emp < 0:
        conclusao.append("os empréstimos diminuíram, mostrando menor dependência de capital de terceiros")
    if dfin < 0:
        conclusao.append("as despesas financeiras caíram em função da menor necessidade de empréstimos")
    if ncg > 0:
        conclusao.append("a NCG aumentou, normalmente porque o ativo operacional cresceu mais que o passivo operacional")
    elif ncg < 0:
        conclusao.append("a NCG diminuiu, reduzindo a necessidade de financiamento da operação")
    if tes > 0:
        conclusao.append("a tesouraria melhorou")
    elif tes < 0:
        conclusao.append("a tesouraria piorou")

    resumo = "; ".join(conclusao) if conclusao else "não foram encontradas variações suficientes para uma conclusão automática forte."
    respostas.append(("Diagnóstico automático pelos valores encontrados",
        f"Pelos valores lidos, o faturamento {sinal(fat)} ({frase_moeda(fat)}), o CPV {sinal(cpv)} ({frase_moeda(cpv)}) e as despesas operacionais {sinal(desp)} ({frase_moeda(desp)}). "
        f"No resultado financeiro, os empréstimos {sinal(emp)} ({frase_moeda(emp)}) e as despesas financeiras {sinal(dfin)} ({frase_moeda(dfin)}). "
        f"No patrimonial, contas a receber {sinal(cr)}, fornecedores {sinal(forn)}, ativo operacional {sinal(ao)}, passivo operacional {sinal(po)}, NCG {sinal(ncg)}, CDG {sinal(cdg)} e tesouraria {sinal(tes)}. "
        f"Conclusão: {resumo}."))
    return respostas


st.title("📊 Interpretador PCO — RF, RC e RP/NCG")
st.write("Envie Excel ou imagem/print da simulação. O app tenta ler original x novo valor e gera resposta no padrão das questões da FAE.")

with st.expander("📚 Base usada pelo interpretador — anotações + PDFs + prints + planilha", expanded=False):
    st.write("Esta versão usa a lógica dos materiais enviados: elaboração de premissas, cenário 1, cenário 2, prints de prova, anotações e estrutura da planilha.")
    for tema, itens in MATERIAL_BASE.items():
        st.markdown(f"**{tema}**")
        for item in itens:
            st.markdown(f"- {item}")

with st.sidebar:
    st.header("Configuração")
    premissa = st.selectbox("Tipo de premissa", [
        "Geral / Roteiro oficial", "Quantidade vendida", "Preço de venda", "Produção / Estoque PA", "Compra / Estoque MP", "Preço dos insumos",
        "PMC / Prazo de recebimento", "PMO / Prazo de pagamento", "Despesas / MKT / Comissão / Transporte", "Pessoal / RH", "Outra"
    ])
    modo = st.radio("Modo", ["Carregar arquivo", "Manual"])
    modo_pergunta = st.radio("Como gerar respostas", ["Banco de perguntas", "Pergunta específica"])
    pergunta_digitada = ""
    if modo_pergunta == "Pergunta específica":
        pergunta_digitada = st.text_area("Cole ou digite a pergunta da prova", height=110, placeholder="Ex.: Por que o Δ Faturamento [%] é superior ao Δ CPV [%]?")

edit = None
texto_ocr = ""

if modo == "Carregar arquivo":
    arq = st.file_uploader("Enviar Excel ou imagem", type=["xlsx", "xlsm", "png", "jpg", "jpeg"])
    if arq:
        nome = arq.name.lower()
        if nome.endswith((".xlsx", ".xlsm")):
            with st.spinner("Lendo planilha..."):
                df = ler_planilha(arq)
        else:
            st.image(arq, caption="Imagem enviada", use_container_width=True)
            with st.spinner("Lendo texto da imagem via OCR..."):
                texto_ocr = ocr_imagem(arq)
            with st.expander("Texto lido da imagem/OCR", expanded=False):
                texto_ocr = st.text_area("Você pode corrigir o texto aqui se o OCR errar", texto_ocr, height=220)
            df = extrair_df_texto(texto_ocr)

        if df.empty:
            st.warning("Não consegui detectar automaticamente as contas. Use o modo manual ou preencha/ajuste a tabela abaixo.")
            df = pd.DataFrame({"Conta": CONTAS_PADRAO, "Original": 0.0, "Novo Valor": 0.0})
        st.subheader("Valores detectados/ajustáveis")
        base = df[[c for c in ["Conta", "Original", "Novo Valor"] if c in df.columns]].copy()
        edit = st.data_editor(base, use_container_width=True, num_rows="dynamic")
    else:
        st.info("Envie a planilha ou imagem para começar.")
else:
    df = pd.DataFrame({"Conta": CONTAS_PADRAO, "Original": 0.0, "Novo Valor": 0.0})
    edit = st.data_editor(df, use_container_width=True, num_rows="dynamic")

if edit is not None:
    edit = montar_df(edit.to_dict("records"))
    st.subheader("Comparativo calculado")
    st.dataframe(edit, use_container_width=True)

    st.subheader("Respostas automáticas")
    respostas = gerar_respostas(edit, premissa)
    for titulo, texto in respostas:
        with st.expander(titulo, expanded=True):
            st.write(texto)

    st.subheader("Texto corrido para colar")
    texto_final = "\n\n".join([f"{t}\n{txt}" for t, txt in respostas])
    st.text_area("Resposta pronta", texto_final, height=360)

    st.download_button(
        "Baixar resposta em TXT",
        data=texto_final.encode("utf-8"),
        file_name="resposta_pco.txt",
        mime="text/plain",
    )
